"""
Shared library for Scifind.
Provides DB connection, LaTeX rendering, dimension formatting,
default unit parsing, and CSV import/export.
"""

import csv
import html
import json
import os
import re
import sqlite3
from collections import defaultdict
from fractions import Fraction
from io import StringIO
from pathlib import Path

DEFAULT_DB = str(Path.home() / ".local" / "share" / "formula" / "formulas.db")

SCIENCES_PATH = Path(__file__).resolve().parent / "sciences.json"

def load_sciences_tree():
    with open(SCIENCES_PATH) as f:
        return json.load(f)["sciences"]


def _name_to_id(name, level="science"):
    """Look up the tree ID for a localized English name, or slugify if not found."""
    sciences = load_sciences_tree()
    if level == "science":
        for s in sciences:
            if s.get("translations", {}).get("en-us") == name:
                return s["id"]
    elif level == "branch":
        for s in sciences:
            for b in s.get("children", []):
                if b.get("translations", {}).get("en-us") == name:
                    return b["id"]
    elif level == "subbranch":
        for s in sciences:
            for b in s.get("children", []):
                for sb in b.get("children", []):
                    if sb.get("translations", {}).get("en-us") == name:
                        return sb["id"]
    elif level == "topic":
        for s in sciences:
            for b in s.get("children", []):
                for c in b.get("children", []):
                    # c could be a subbranch (with children) or a direct topic (no children)
                    if isinstance(c, dict) and "children" in c and c["children"]:
                        for t in c["children"]:
                            if t.get("translations", {}).get("en-us") == name:
                                return t["id"]
                    else:
                        if c.get("translations", {}).get("en-us") == name:
                            return c["id"]
    return _slugify(name)


def _slugify(text):
    return text.lower().replace(" ", "_").replace("-", "_").replace("'", "").replace("/", "_per_")


def save_sciences_tree(sciences):
    with open(SCIENCES_PATH, "w") as f:
        json.dump({"sciences": sciences}, f, indent=2)
        f.write("\n")


def sync_sciences_from_db(conn):
    """Scan formulas and quantities for science/branch/subbranch/topic values missing from sciences.json.
    Works with both JSON i18n columns and plain ID columns.
    """
    def _get_vals(table, col):
        rows = conn.execute(f"""
            SELECT DISTINCT CASE WHEN {col} LIKE '{{%' THEN json_extract({col}, '$.en-us') ELSE {col} END AS val
            FROM {table}
            WHERE {col} IS NOT NULL AND {col} != ''
        """).fetchall()
        return {r["val"] for r in rows if r["val"]}

    science_names = _get_vals("(SELECT science FROM formula UNION ALL SELECT science FROM quantity)", "science")
    branch_names = _get_vals("(SELECT branch FROM formula UNION ALL SELECT branch FROM quantity)", "branch")
    subbranch_names = _get_vals("(SELECT subbranch FROM formula UNION ALL SELECT subbranch FROM quantity)", "subbranch")
    topic_names = _get_vals("(SELECT topic FROM formula UNION ALL SELECT topic FROM quantity)", "topic")

    sciences = load_sciences_tree()
    known_science_ids = {s["id"] for s in sciences}
    known_branch_ids = set()
    known_subbranch_ids = set()
    known_topic_ids = set()

    # Build maps and collect known IDs
    science_name_to_id = {}
    for s in sciences:
        en = s.get("translations", {}).get("en-us", "")
        if en:
            science_name_to_id[en] = s["id"]
        for b in s.get("children", []):
            known_branch_ids.add(b["id"])
            for c in b.get("children", []):
                if isinstance(c, dict) and "children" in c and c["children"]:
                    known_subbranch_ids.add(c["id"])
                    for t in c["children"]:
                        known_topic_ids.add(t["id"])
                elif isinstance(c, dict):
                    known_topic_ids.add(c["id"])

    changed = False

    # Add missing sciences
    for name in science_names:
        if name not in science_name_to_id:
            sid = _slugify(name)
            if sid not in known_science_ids:
                sciences.append({"id": sid, "translations": {"en-us": name}, "children": []})
                known_science_ids.add(sid)
                changed = True

    # Add missing branches
    branch_rows = conn.execute("""
        SELECT DISTINCT CASE WHEN branch LIKE '{%' THEN json_extract(branch, '$.en-us') ELSE branch END AS bname,
                        CASE WHEN science LIKE '{%' THEN json_extract(science, '$.en-us') ELSE science END AS sname
        FROM (SELECT branch, science FROM formula WHERE branch IS NOT NULL AND branch != '' AND science IS NOT NULL AND science != ''
              UNION ALL
              SELECT branch, science FROM quantity WHERE branch IS NOT NULL AND branch != '' AND science IS NOT NULL AND science != '')
    """).fetchall()

    for row in branch_rows:
        bname = row["bname"]
        sname = row["sname"]
        if not bname or not sname:
            continue
        bid = _slugify(bname)
        if bid in known_branch_ids:
            continue
        sid = science_name_to_id.get(sname, _slugify(sname))
        sci_node = next((s for s in sciences if s["id"] == sid), None)
        if sci_node is None:
            sci_node = {"id": sid, "translations": {"en-us": sname}, "children": []}
            sciences.append(sci_node)
            known_science_ids.add(sid)
        existing = [b for b in sci_node.get("children", []) if b["id"] == bid]
        if not existing:
            sci_node.setdefault("children", []).append({
                "id": bid, "translations": {"en-us": bname}, "children": []
            })
            known_branch_ids.add(bid)
            changed = True

    # Add missing subbranches — scan DB for subbranch+branch pairs
    sb_rows = conn.execute("""
        SELECT DISTINCT CASE WHEN subbranch LIKE '{%' THEN json_extract(subbranch, '$.en-us') ELSE subbranch END AS sbname,
                        CASE WHEN branch LIKE '{%' THEN json_extract(branch, '$.en-us') ELSE branch END AS bname
        FROM (SELECT subbranch, branch FROM formula WHERE subbranch IS NOT NULL AND subbranch != '' AND branch IS NOT NULL AND branch != ''
              UNION ALL
              SELECT subbranch, branch FROM quantity WHERE subbranch IS NOT NULL AND subbranch != '' AND branch IS NOT NULL AND branch != '')
    """).fetchall()

    for row in sb_rows:
        sbname = row["sbname"]
        bname = row["bname"]
        if not sbname or not bname:
            continue
        sbid = _slugify(sbname)
        if sbid in known_subbranch_ids:
            continue
        bid = _slugify(bname)
        for s in sciences:
            for b in s.get("children", []):
                if b["id"] == bid:
                    existing = [c for c in b.get("children", []) if c["id"] == sbid]
                    if not existing:
                        b.setdefault("children", []).append({
                            "id": sbid, "translations": {"en-us": sbname}, "children": []
                        })
                        known_subbranch_ids.add(sbid)
                        changed = True
                    break

    # Add missing topics — can be under a branch (direct) or under a subbranch
    topic_rows = conn.execute("""
        SELECT DISTINCT CASE WHEN topic LIKE '{%' THEN json_extract(topic, '$.en-us') ELSE topic END AS tname,
                        CASE WHEN branch LIKE '{%' THEN json_extract(branch, '$.en-us') ELSE branch END AS bname,
                        CASE WHEN subbranch LIKE '{%' THEN json_extract(subbranch, '$.en-us') ELSE subbranch END AS sbname
        FROM (SELECT topic, branch, subbranch FROM formula WHERE topic IS NOT NULL AND topic != '' AND branch IS NOT NULL AND branch != ''
              UNION ALL
              SELECT topic, branch, subbranch FROM quantity WHERE topic IS NOT NULL AND topic != '' AND branch IS NOT NULL AND branch != '')
    """).fetchall()

    for row in topic_rows:
        tname = row["tname"]
        bname = row["bname"]
        sbname = row["sbname"]
        if not tname or not bname:
            continue
        tid = _slugify(tname)
        if tid in known_topic_ids:
            continue
        bid = _slugify(bname)
        for s in sciences:
            for b in s.get("children", []):
                if b["id"] != bid:
                    continue
                if sbname:
                    sbid = _slugify(sbname)
                    for sb in b.get("children", []):
                        if sb["id"] == sbid:
                            existing = [t for t in sb.get("children", []) if t["id"] == tid]
                            if not existing:
                                sb.setdefault("children", []).append({"id": tid, "translations": {"en-us": tname}})
                                known_topic_ids.add(tid)
                                changed = True
                            break
                else:
                    existing = [t for t in b.get("children", []) if t["id"] == tid]
                    if not existing:
                        b.setdefault("children", []).append({"id": tid, "translations": {"en-us": tname}})
                        known_topic_ids.add(tid)
                        changed = True
                break

    # Prune unused topics, subbranches, and branches from the tree
    ref_topics = set()
    for row in conn.execute("""
        SELECT topic FROM formula WHERE topic IS NOT NULL AND topic != '' AND topic NOT LIKE '{%'
        UNION ALL
        SELECT topic FROM quantity WHERE topic IS NOT NULL AND topic != '' AND topic NOT LIKE '{%'
    """).fetchall():
        ref_topics.add(row["topic"])
    for row in conn.execute("""
        SELECT json_extract(topic, '$.en-us') AS val FROM formula WHERE topic LIKE '{%'
        UNION ALL
        SELECT json_extract(topic, '$.en-us') AS val FROM quantity WHERE topic LIKE '{%'
    """).fetchall():
        v = row["val"]
        if v:
            ref_topics.add(_slugify(v))

    for s in sciences[:]:
        for b in list(s.get("children", [])):
            for c in list(b.get("children", [])):
                if isinstance(c, dict) and "children" in c and c["children"]:
                    c["children"] = [t for t in c["children"] if t["id"] in ref_topics]
                    if not c["children"]:
                        b["children"].remove(c)
                        changed = True
                elif isinstance(c, dict):
                    if c["id"] not in ref_topics:
                        b["children"].remove(c)
                        changed = True
            if not b.get("children"):
                s["children"].remove(b)
                changed = True
        if not s.get("children"):
            sciences.remove(s)
            changed = True

    if changed:
        save_sciences_tree(sciences)

    return sciences

def tl(translations, locale):
    """Translation lookup from a translations dict."""
    if not translations:
        return ""
    if isinstance(translations, str):
        try:
            translations = json.loads(translations)
        except (json.JSONDecodeError, TypeError):
            return translations
    return translations.get(locale, translations.get("en-us", ""))

DIM_ORDER = ["M", "L", "T", "I", "\u0398", "N", "J"]
DIM_COLS = ["dim_M", "dim_L", "dim_T", "dim_I", "dim_\u0398", "dim_N", "dim_J"]
def get_dim_var_ids(conn):
    """Build {dim_symbol: quantity_id} from base SI quantities.

    Queries the DB for quantities with is_dim = 1 and finds which
    dimension column (dim_M, dim_L, etc.) is non-zero for each.
    For robustness, each dim column is matched separately so that
    data errors (e.g. two quantities with dim_M=1) don't collide.
    """
    DIM_SYMS = ["M", "L", "T", "I", "\u0398", "N", "J"]
    DIM_COLS = ["dim_M", "dim_L", "dim_T", "dim_I", "dim_\u0398", "dim_N", "dim_J"]
    result = {}
    for sym, col in zip(DIM_SYMS, DIM_COLS):
        row = conn.execute(
            f"SELECT id FROM quantity WHERE is_dim = 1 AND {col} != 0 LIMIT 1"
        ).fetchone()
        if row:
            result[sym] = row["id"]
    return result


def db_path():
    return os.environ.get("FORMULA_DB", DEFAULT_DB)


def get_conn():
    p = db_path()
    os.makedirs(os.path.dirname(p), exist_ok=True)
    conn = sqlite3.connect(p)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("PRAGMA synchronous = NORMAL")
    return conn


def en(data, locale="en-us"):
    if not data:
        return ""
    try:
        d = json.loads(data)
        for key in (locale, "en-us"):
            val = d.get(key)
            if val:
                return val
        return str(data)
    except (json.JSONDecodeError, TypeError):
        return str(data)


def rebuild_fts(conn):
    conn.execute("DELETE FROM formula_fts")
    rows = conn.execute("""
        SELECT f.id, f.name, f.description,
               COALESCE(group_concat(q.name_en, ' '), '') AS vars
        FROM formula f
        LEFT JOIN formula_item fi ON fi.formula_id = f.id
        LEFT JOIN (SELECT DISTINCT id, json_extract(name, '$.en-us') AS name_en
                   FROM quantity) q ON q.id = fi.quantity_id
        GROUP BY f.id
    """).fetchall()
    for r in rows:
        conn.execute(
            "INSERT INTO formula_fts (formula_id, name, description, quantities) VALUES (?, ?, ?, ?)",
            (r["id"], r["name"], r["description"], r["vars"]),
        )

    conn.execute("DELETE FROM quantity_fts")
    vrows = conn.execute("""
        SELECT id, json_extract(name, '$.en-us') AS name_en, symbol
        FROM quantity
    """).fetchall()
    for r in vrows:
        conn.execute(
            "INSERT INTO quantity_fts (quantity_id, name, symbol) VALUES (?, ?, ?)",
            (r["id"], r["name_en"], r["symbol"]),
        )

    conn.execute("DELETE FROM unit_fts")
    urows = conn.execute("""
        SELECT id, json_extract(name, '$.en-us') AS name_en, symbol
        FROM unit
    """).fetchall()
    for r in urows:
        conn.execute(
            "INSERT INTO unit_fts (unit_id, name, symbol) VALUES (?, ?, ?)",
            (r["id"], r["name_en"], r["symbol"]),
        )

    conn.commit()
    return len(rows)


def migrate_db(conn):
    """Apply schema migrations for existing databases."""
    # Add subbranch column if missing
    for table in ("formula", "quantity"):
        tcols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        if "subbranch" not in tcols:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN subbranch TEXT")
    conn.commit()

    cols = [r[1] for r in conn.execute("PRAGMA table_info(unit)").fetchall()]
    if "name" not in cols:
        conn.execute("ALTER TABLE unit ADD COLUMN name TEXT NOT NULL DEFAULT '{}'")
        for row in conn.execute("SELECT id FROM unit").fetchall():
            uid = row["id"]
            en_name = uid.replace("_", " ").title()
            conn.execute("UPDATE unit SET name = ? WHERE id = ?",
                         (json.dumps({"en": en_name}), uid))
        conn.commit()

    # Migrate old JSON i18n {"en":"..."} to {"en-us":"..."}
    for table in ("formula", "quantity", "unit", "condition"):
        tcols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        json_cols = []
        for c in ("name", "science", "branch", "topic", "description"):
            if c in tcols:
                json_cols.append(c)
        for col in json_cols:
            # Only process if column still has JSON values
            conn.execute(f"""
                UPDATE {table}
                SET {col} = json_set({col}, '$.en-us', json_extract({col}, '$.en'))
                WHERE {col} LIKE '{{%'
                  AND json_extract({col}, '$.en') IS NOT NULL
                  AND json_extract({col}, '$.en-us') IS NULL
            """)
            conn.execute(f"""
                UPDATE {table}
                SET {col} = json_remove({col}, '$.en')
                WHERE {col} LIKE '{{%'
                  AND json_extract({col}, '$.en-us') IS NOT NULL
                  AND json_extract({col}, '$.en') IS NOT NULL
            """)
    conn.commit()

    # Add en-uk locale keys for known US/UK spelling differences
    UK_OVERRIDES = {
        "Meter": "Metre", "Centimeter": "Centimetre", "Decimeter": "Decimetre",
        "Liter": "Litre", "Center of Mass": "Centre of Mass",
    }
    for row in conn.execute("SELECT id, name FROM unit WHERE json_extract(name, '$.en-uk') IS NULL").fetchall():
        try:
            d = json.loads(row["name"])
            en_us = d.get("en-us", "")
            for us_spelling, uk_spelling in UK_OVERRIDES.items():
                if us_spelling in en_us:
                    uk_val = en_us.replace(us_spelling, uk_spelling)
                    if uk_val != en_us:
                        conn.execute(
                            "UPDATE unit SET name = json_set(name, '$.en-uk', ?) WHERE id = ?",
                            (uk_val, row["id"]),
                        )
                        break
        except (json.JSONDecodeError, TypeError):
            pass
    conn.commit()

    # Migrate old-style LaTeX symbols to siunitx-compatible ones
    conn.execute(
        "UPDATE unit SET symbol = '\\degreeCelsius' "
        "WHERE symbol = '\\\\text{\\\\textdegree C}'"
    )
    conn.execute("UPDATE unit SET symbol = '\\ohm' WHERE symbol = '\\\\Omega'")
    conn.execute(
        "UPDATE unit SET symbol = '\\newton' "
        "WHERE symbol = 'N' AND quantity_id = 'force'"
    )
    conn.commit()

    # Normalize double backslashes in symbols: \\command → \command
    conn.execute("UPDATE unit SET symbol = substr(symbol, 2) WHERE symbol LIKE '\\\\%' AND symbol NOT LIKE '\\\\mathrm%'")
    conn.commit()

    # Ensure FTS tables exist
    for fts_sql in [
        "CREATE VIRTUAL TABLE IF NOT EXISTS quantity_fts USING fts5(quantity_id UNINDEXED, name, symbol)",
        "CREATE VIRTUAL TABLE IF NOT EXISTS unit_fts USING fts5(unit_id UNINDEXED, name, symbol)",
    ]:
        try:
            conn.execute(fts_sql)
        except Exception:
            pass
    conn.commit()

    # Migrate science/branch/topic from JSON i18n to simple IDs
    for table in ("formula", "quantity"):
        tcols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        for col in ("science", "branch", "subbranch", "topic"):
            if col not in tcols:
                continue
            # Check if values are still JSON (start with '{')
            sample = conn.execute(f"SELECT {col} FROM {table} WHERE {col} LIKE '{{%' LIMIT 1").fetchone()
            if sample and sample[0]:
                # First, sync sciences.json so all names have IDs
                sync_sciences_from_db(conn)
                # Convert: extract English name, look up tree ID, update
                cur = conn.execute(f"SELECT rowid, json_extract({col}, '$.en-us') AS val FROM {table} WHERE {col} LIKE '{{%'")
                updates = []
                for row in cur.fetchall():
                    en_name = row["val"]
                    if en_name:
                        tid = _name_to_id(en_name, col)
                        updates.append((tid, row["rowid"]))
                for tid, rowid in updates:
                    conn.execute(f"UPDATE {table} SET {col} = ? WHERE rowid = ?", (tid, rowid))
                conn.commit()

    # Add symbol_overwrite column to quantity if missing
    qcols = [r[1] for r in conn.execute("PRAGMA table_info(quantity)").fetchall()]
    if "symbol_overwrite" not in qcols:
        conn.execute("ALTER TABLE quantity ADD COLUMN symbol_overwrite TEXT")
        conn.commit()

    # Add quantity_name_overwrite column to formula_item if missing
    ficols = [r[1] for r in conn.execute("PRAGMA table_info(formula_item)").fetchall()]
    if "quantity_name_overwrite" not in ficols:
        conn.execute("ALTER TABLE formula_item ADD COLUMN quantity_name_overwrite TEXT")
        conn.commit()

    # Add latex_coef column to formula_item (replaces coeff_special)
    ficols = [r[1] for r in conn.execute("PRAGMA table_info(formula_item)").fetchall()]
    if "latex_coef" not in ficols:
        conn.execute("ALTER TABLE formula_item ADD COLUMN latex_coef TEXT")
        conn.execute("UPDATE formula_item SET latex_coef = CASE coeff_special WHEN 'pi' THEN '\\pi' WHEN 'e' THEN 'e' WHEN 'a' THEN 'a' WHEN 'b' THEN 'b' ELSE coeff_special END WHERE coeff_special IS NOT NULL")
        conn.commit()

    # Add latex_factor column to unit if missing
    ucols = [r[1] for r in conn.execute("PRAGMA table_info(unit)").fetchall()]
    if "latex_factor" not in ucols:
        conn.execute("ALTER TABLE unit ADD COLUMN latex_factor TEXT")
        conn.commit()

    # Rename base_si → is_dim in quantity table
    qcols = [r[1] for r in conn.execute("PRAGMA table_info(quantity)").fetchall()]
    if "base_si" in qcols and "is_dim" not in qcols:
        conn.execute("ALTER TABLE quantity ADD COLUMN is_dim INTEGER NOT NULL DEFAULT 0")
        conn.execute("UPDATE quantity SET is_dim = base_si")
        conn.commit()

    # Ensure the 7 SI base quantities are marked as is_dim = 1
    base_ids = ('mass', 'length', 'time', 'current', 'temperature', 'amount', 'luminous_intensity')
    conn.execute(
        "UPDATE quantity SET is_dim = 1 WHERE id IN (?,?,?,?,?,?,?) AND is_dim = 0",
        base_ids
    )
    conn.commit()


def fmt_num(n):
    if n == int(n):
        return str(int(n))
    s = f"{n:.10f}".rstrip("0")
    return s.rstrip(".")


# Dimension rendering

def render_dimensions(dim_M=0, dim_L=0, dim_T=0, dim_I=0, dim_Theta=0, dim_N=0, dim_J=0):
    """Render dimension exponents as human-readable string."""
    vals = [dim_M, dim_L, dim_T, dim_I, dim_Theta, dim_N, dim_J]
    parts = []
    for k, exp in zip(DIM_ORDER, vals):
        if exp is None:
            exp = 0
        if exp != 0:
            p = k
            if exp != 1:
                p += f"^{fmt_num(exp)}"
            parts.append(p)
    return " \u00b7 ".join(parts) if parts else "dimensionless"


def render_dimensions_latex(dim_M=0, dim_L=0, dim_T=0, dim_I=0, dim_Theta=0, dim_N=0, dim_J=0,
                            var_latex_map=None, unit_symbol_map=None, dim_latex_map=None, mode="var"):
    """Render dimension exponents as LaTeX.

    var_latex_map:  dict mapping dim_sym (M,L,T,I,Θ,N,J) to quantity-symbol LaTeX
    unit_symbol_map: dict mapping dim_sym (M,L,T,I,Θ,N,J) to unit-symbol LaTeX
    dim_latex_map:   dict mapping dim_sym (M,L,T,I,Θ,N,J) to dimension-symbol LaTeX (from symbol_overwrite['dim'])
    mode: "dim", "var" (default), or "unit"
    """
    vals = [dim_M, dim_L, dim_T, dim_I, dim_Theta, dim_N, dim_J]
    if mode == "dim" and dim_latex_map:
        lookup = dim_latex_map
    elif mode == "var" and var_latex_map:
        lookup = var_latex_map
    else:
        lookup = unit_symbol_map
    parts = []
    for dim_sym, exp in zip(DIM_ORDER, vals):
        if not exp:
            continue
        sym = lookup.get(dim_sym) if lookup else None
        if sym is None:
            sym = f"\\mathrm{{{dim_sym}}}"
        if exp == 1:
            parts.append(sym)
        else:
            e = str(int(exp)) if exp == int(exp) else str(exp)
            parts.append(f"{sym}^{{{e}}}")
    if not parts:
        return "\\text{dimensionless}"
    return " \\cdot ".join(parts)


def dims_from_row(row):
    """Extract dimension values from a row (dict or sqlite3.Row)."""
    cols = ["dim_M", "dim_L", "dim_T", "dim_I", "dim_\u0398", "dim_N", "dim_J"]
    return [dict(row).get(c, 0) or 0 for c in cols]


# Default unit rendering

def parse_default_unit_json(default_unit_json):
    """Parse default_unit JSON and return list of (unit_id, exponent)."""
    if not default_unit_json:
        return []
    try:
        parts = json.loads(default_unit_json)
        return [(p["unit"], p["exponent"]) for p in parts]
    except (json.JSONDecodeError, KeyError, TypeError):
        return []


def decompose_default_unit_parts(default_unit_json):
    """Parse default_unit JSON and decompose composite units (e.g. metre_per_second) into base parts.

    Returns list of (unit_id, exponent) with composites expanded.
    """
    parts = parse_default_unit_json(default_unit_json)
    result = []
    for uid, exp in parts:
        comp = unit_id_components(uid)
        if comp is None:
            result.append((uid, exp))
        else:
            nums, dens = comp
            for nid, nexp in nums:
                result.append((nid, nexp * exp))
            if dens:
                for did, dexp in dens:
                    result.append((did, -dexp * exp))
    return result


def render_default_unit_html(default_unit_json, unit_url_func=None, unit_name_func=None, locale="en-us"):
    """Render default_unit JSON as HTML with links."""
    parts = decompose_default_unit_parts(default_unit_json)
    if not parts:
        return ""
    ls = _get_ls(locale)
    num_parts = []
    den_parts = []
    for uid, exp in parts:
        if exp >= 0:
            num_parts.append((uid, exp))
        else:
            den_parts.append((uid, -exp))
    num_str = _render_unit_group(num_parts, unit_url_func, unit_name_func, locale)
    if not den_parts:
        return num_str
    den_str = _render_unit_group(den_parts, unit_url_func, unit_name_func, locale)
    if not num_str:
        return f"{ls['reciprocal']} {den_str}"
    return f"{num_str} {ls['per']} {den_str}"


def render_default_unit_symbol(default_unit_json, unit_symbol_func=None):
    """Render default_unit JSON as LaTeX symbol expression (e.g. \\mathrm{m}\\cdot\\mathrm{s}^{-2})."""
    parts = decompose_default_unit_parts(default_unit_json)
    if not parts:
        return ""
    num_parts = []
    den_parts = []
    for uid, exp in parts:
        if exp >= 0:
            num_parts.append((uid, exp))
        else:
            den_parts.append((uid, -exp))
    def _render(parts):
        items = []
        for uid, exp in parts:
            sym = unit_symbol_func(uid) if unit_symbol_func else uid
            items.append(sym + (f"^{{{int(exp)}}}" if exp != 1 else ""))
        return " \\cdot ".join(items) if items else ""
    num_str = _render(num_parts)
    den_str = _render(den_parts)
    if not den_str:
        return num_str
    if not num_str:
        return f"1 / ({den_str})" if len(den_parts) > 1 else f"1 / {den_str}"
    return f"{num_str} / ({den_str})" if len(den_parts) > 1 else f"{num_str} / {den_str}"


def _render_unit_group(parts, url_func, name_func=None, locale="en-us"):
    """Render a list of (unit_id, exponent) as HTML with natural-language exponents."""
    items = []
    for uid, exp in parts:
        label = name_func(uid) if name_func else uid.replace("_", " ").title()
        word = _exp_word(exp, locale)
        if url_func:
            link = url_func(uid)
            text = f'<a href="{html.escape(link)}">{html.escape(label)}</a>'
        else:
            text = html.escape(label)
        if word:
            text += " " + html.escape(word)
        items.append(text)
    return "-".join(items)


# Unit ID decomposition

def is_composite_unit(unit_id):
    """Return True if the unit ID describes a derived/composite unit."""
    return (
        "_per_" in unit_id
        or unit_id.startswith("square_")
        or unit_id.startswith("cubic_")
        or unit_id.startswith("reciprocal_")
    )

def _strip_unit_suffix(uid):
    """Remove disambiguation suffixes from unit IDs (e.g. _mass, _energy)."""
    for suffix in ("_mass", "_energy", "_luminous", "_mechanical", "_thermal",
                   "_gravitational", "_molar", "_specific", "_radiant"):
        if uid.endswith(suffix):
            return uid[:-len(suffix)]
    return uid


def _parse_unit_part(part_str):
    """Parse a numerator/denominator part string into [(unit_id, exponent)]."""
    SUFFIXES = ["_sq"] + [f"_{s}" for s in
        ("mass", "energy", "luminous", "mechanical", "thermal",
         "gravitational", "molar", "specific", "radiant")]
    markers = {}
    for i, sfx in enumerate(SUFFIXES):
        marker = f"\x00{i}"
        markers[marker] = sfx
        if part_str.endswith(sfx):
            part_str = part_str[:-len(sfx)] + marker

    tokens = part_str.split("_")
    result = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        if not tok:
            i += 1
            continue
        exp = 1
        if tok == "square":
            exp = 2
            i += 1
            if i < len(tokens):
                tok = tokens[i]
            else:
                break
        elif tok == "cubic":
            exp = 3
            i += 1
            if i < len(tokens):
                tok = tokens[i]
            else:
                break
        elif tok == "reciprocal":
            exp = -1
            i += 1
            if i < len(tokens):
                tok = tokens[i]
            else:
                break
        for marker, orig in sorted(markers.items(), key=lambda x: -len(x[0])):
            if tok == marker:
                tok = ""
                break
            if tok.endswith(marker):
                tok = tok[: -len(marker)]
                if orig == "_sq":
                    exp = 2
                break
        if not tok:
            i += 1
            continue
        tok = _strip_unit_suffix(tok)
        result.append((tok, exp))
        i += 1
    return result


def unit_id_components(unit_id):
    """Parse a composite unit ID into (num_parts, den_parts).

    Returns (num_units, den_units) or None for a simple (non-composite) unit.
    Each part is a list of (base_unit_id, exponent).
    """
    parts = unit_id.split("_per_")
    if len(parts) == 1:
        pid = parts[0]
        if any(pid.startswith(p) for p in ("square_", "cubic_", "reciprocal_")):
            nums = _parse_unit_part(pid)
            return (nums, None)
        return None  # Simple unit

    num_str = parts[0]
    den_str = "_per_".join(parts[1:])
    nums = _parse_unit_part(num_str)
    dens = _parse_unit_part(den_str)
    return (nums, dens)


LOCALE_STRINGS = {
    "en-us": {
        "squared": "squared",
        "cubed": "cubed",
        "inverse": "inverse",
        "to_the": "to the",
        "per": "per",
        "reciprocal": "Reciprocal",
    },
    "en-uk": {
        "squared": "squared",
        "cubed": "cubed",
        "inverse": "inverse",
        "to_the": "to the",
        "per": "per",
        "reciprocal": "Reciprocal",
    },
}

ORDINAL_SPECIAL = {11: "th", 12: "th", 13: "th"}
ORDINAL_LAST_DIGIT = {0: "th", 1: "st", 2: "nd", 3: "rd",
                      4: "th", 5: "th", 6: "th", 7: "th", 8: "th", 9: "th"}


def difficulty_stars(difficulty, max_dots=5):
    """Render a difficulty score (1-10) as filled + empty stars."""
    d = min(int(difficulty or 0), max_dots)
    return "\u2605" * d + "\u2606" * (max_dots - d)


def _get_ls(locale):
    """Get locale strings, falling back to en-us."""
    return LOCALE_STRINGS.get(locale, LOCALE_STRINGS["en-us"])


def _ordinal(n):
    last2 = n % 100
    if last2 in ORDINAL_SPECIAL:
        suffix = ORDINAL_SPECIAL[last2]
    else:
        suffix = ORDINAL_LAST_DIGIT[n % 10]
    return f"{n}{suffix}"


def _exp_word(exp, locale="en-us"):
    """Return natural-language exponent word for a unit."""
    ls = _get_ls(locale)
    if exp == 2:
        return ls["squared"]
    elif exp == 3:
        return ls["cubed"]
    elif exp == 1:
        return ""
    elif exp == -1:
        return ls["inverse"]
    elif exp > 3:
        return f"{ls['to_the']} {_ordinal(exp)}"
    else:
        return ""


def render_unit_decomposition(unit_id, name_func, url_func=None, locale="en-us"):
    """Render a unit name with links, decomposing composite units into components."""
    comp = unit_id_components(unit_id)
    if comp is None:
        name = name_func(unit_id) if name_func else unit_id.replace("_", " ").title()
        if url_func:
            url = url_func(unit_id)
            return f'<a href="{html.escape(url)}">{html.escape(name)}</a>'
        return html.escape(name)

    nums, dens = comp

    fixed_nums = []
    fixed_dens = list(dens) if dens else []
    for uid, exp in nums:
        if exp < 0:
            fixed_dens.append((uid, -exp))
        else:
            fixed_nums.append((uid, exp))

    ls = _get_ls(locale)
    if fixed_nums and fixed_dens:
        num_html = _render_unit_group(fixed_nums, url_func, name_func, locale)
        den_html = _render_unit_group(fixed_dens, url_func, name_func, locale)
        return f"{num_html} {ls['per']} {den_html}"
    elif fixed_dens:
        den_html = _render_unit_group(fixed_dens, url_func, name_func, locale)
        return f"{ls['reciprocal']} {den_html}"
    return _render_unit_group(fixed_nums, url_func, name_func, locale)


# LaTeX rendering helpers

def _is_integer(n):
    if n is None:
        return False
    return n == int(n)


def _render_exp(exp, base_str):
    """Render a variable/coeff part with exponent handling."""
    if exp is None or exp == 1:
        return base_str

    if _is_integer(exp):
        iexp = int(exp)
        if iexp < 0:
            inner = base_str
            if iexp != -1:
                inner += "^{" + fmt_num(-iexp) + "}"
            return "\\frac{1}{" + inner + "}"
        return base_str + "^{" + fmt_num(iexp) + "}"

    try:
        f = Fraction(exp).limit_denominator(100)
        num = f.numerator
        den = f.denominator
    except Exception:
        return base_str + "^{" + fmt_num(exp) + "}"

    if num == 1:
        if den == 2:
            return "\\sqrt{" + base_str + "}"
        return "\\sqrt[" + str(den) + "]{" + base_str + "}"
    if num == -1:
        if den == 2:
            return "\\frac{1}{\\sqrt{" + base_str + "}}"
        return "\\frac{1}{\\sqrt[" + str(den) + "]{" + base_str + "}}"
    inner = base_str + "^{" + fmt_num(num) + "}"
    if den == 2:
        return "\\sqrt{" + inner + "}"
    return "\\sqrt[" + str(den) + "]{" + inner + "}"


def _resolve_i18n(value, locale):
    """Resolve a value that may be JSON i18n or plain text."""
    if not value:
        return value
    s = value.strip()
    if s.startswith("{"):
        try:
            d = json.loads(s)
            return d.get(locale, d.get("en-us", s))
        except (json.JSONDecodeError, TypeError):
            return s
    return s


def render_variable(item, flipped, locale="en-us"):
    raw_overwrite = item["symbol_overwrite"] or ""
    var = _resolve_i18n(raw_overwrite, locale) or item["quantity_symbol"] or item["quantity_id"] or "?"
    prefix = item["latex_prefix"] or ""
    suffix = item["latex_suffix"] or ""

    if prefix:
        # Add {} separator if prefix ends with a LaTeX command (backslash + letters)
        # and the variable starts with an alphabetic character, to prevent LaTeX
        # from interpreting the variable as part of the command name.
        if prefix and var and prefix[-1].isalpha() and var[0].isalpha():
            sep = "{}"
        else:
            sep = ""
        var = prefix + sep + var
    if suffix:
        if var and suffix and var[-1].isalpha() and suffix[0].isalpha():
            var = var + "{}" + suffix
        else:
            var = var + suffix

    raw_label = item["label"] or ""
    label = _resolve_i18n(raw_label, locale)
    if label and "_" not in var:
        var += "_{" + label + "}"

    exp = item["var_exponent"] if item["var_exponent"] is not None else 1
    if flipped:
        exp = -exp

    return _render_exp(exp, var)


def render_coeff(item, first_in_term):
    body = None
    is_neg = False
    latex_coef = item["latex_coef"] or None
    coeff_value = item["coeff_value"]
    coeff_exp = item["coeff_exponent"] if item["coeff_exponent"] is not None else 1

    if latex_coef:
        body = latex_coef
    elif coeff_value is not None:
        v = coeff_value
        if v < 0:
            is_neg = True
            v = abs(v)
        body = fmt_num(v) if v != 1 else None
    else:
        body = None

    if body is not None:
        body = _render_exp(coeff_exp, body)
        if not latex_coef and body == "1":
            body = None
    elif coeff_exp is not None and coeff_exp != 1:
        body = _render_exp(coeff_exp, "1")

    if first_in_term and is_neg and body in (None, "", "1"):
        body = ""
        is_neg = False

    return body if body else None, is_neg


def _join_parts(parts):
    """Join LaTeX fragments, adding {} between a command ending and a following letter."""
    result = []
    for p in parts:
        if result and result[-1] and p:
            last = result[-1]
            last_end_cmd = False
            # Check if last part ends with a command (backslash + letters)
            i = len(last) - 1
            while i >= 0 and last[i].isalpha():
                i -= 1
            if i >= 0 and last[i] == '\\' and i < len(last) - 1:
                last_end_cmd = True
            if last_end_cmd and p[0].isalpha():
                result.append("{}")
        result.append(p)
    return "".join(result)


def _render_items_group(items, flipped, locale="en-us"):
    """Render a list of items, returning (numerator_str, denominator_str)."""
    num_parts = []
    den_parts = []
    for i, item in enumerate(items):
        c, neg = render_coeff(item, first_in_term=(i == 0))

        if item["quantity_id"]:
            exp = item["var_exponent"] if item["var_exponent"] is not None else 1
            if flipped:
                exp = -exp
            if neg and c:
                c = "-" + c
            if exp < 0:
                v = render_variable(
                    dict(item, var_exponent=abs(exp)),
                    flipped=False,
                    locale=locale,
                )
                if c:
                    den_parts.append(c)
                den_parts.append(v)
            else:
                v = render_variable(item, flipped=flipped, locale=locale)
                if c:
                    num_parts.append(c)
                num_parts.append(v)
        elif c:
            if neg:
                den_parts.append(c)
            else:
                num_parts.append(c)

    return _join_parts(num_parts), _join_parts(den_parts)


def render_formula_items(items, locale="en-us"):
    """Build LaTeX string from formula_item rows. Handles fractions for negatives."""
    by_term = defaultdict(list)
    for item in items:
        by_term[item["term"]].append(item)

    lhs_terms = []
    rhs_terms = []

    for term_num in sorted(by_term):
        term_items = sorted(by_term[term_num], key=lambda x: (not x["is_primary"], x["sort_order"]))
        primary = [i for i in term_items if i["is_primary"]]
        non_primary = [i for i in term_items if not i["is_primary"]]

        lhs_num, lhs_den = _render_items_group(primary, flipped=True, locale=locale)
        if lhs_num or lhs_den:
            if lhs_den:
                has_neg = lhs_num.startswith("-")
                clean_num = lhs_num.lstrip("-")
                if has_neg:
                    lhs_str = f"-\\frac{{{clean_num}}}{{{lhs_den}}}"
                else:
                    lhs_str = f"\\frac{{{lhs_num}}}{{{lhs_den}}}" if lhs_num else f"\\frac{{1}}{{{lhs_den}}}"
                if lhs_str == "\\frac{}{}":
                    lhs_str = "\\frac{1}{" + lhs_den + "}"
            else:
                lhs_str = lhs_num
            if lhs_str:
                lhs_terms.append(lhs_str)

        rhs_num, rhs_den = _render_items_group(non_primary, flipped=False, locale=locale)
        if rhs_num or rhs_den:
            if rhs_den:
                has_neg = rhs_num.startswith("-")
                clean_num = rhs_num.lstrip("-")
                if has_neg:
                    rhs_str = f"-\\frac{{{clean_num}}}{{{rhs_den}}}"
                else:
                    rhs_str = f"\\frac{{{rhs_num}}}{{{rhs_den}}}" if rhs_num else f"\\frac{{1}}{{{rhs_den}}}"
                if rhs_str == "\\frac{}{}":
                    rhs_str = "\\frac{1}{" + rhs_den + "}"
            else:
                rhs_str = rhs_num
            if rhs_str:
                sign = "+"
                for item in non_primary:
                    cv = item["coeff_value"]
                    if cv is not None and cv < 0:
                        sign = "-"
                        break
                rhs_terms.append((sign, rhs_str))

    lhs = " + ".join(lhs_terms) if lhs_terms else "1"

    rhs_parts = []
    for i, (sign, term_str) in enumerate(rhs_terms):
        if i == 0:
            rhs_parts.append(f"- {term_str}" if sign == "-" else term_str)
        else:
            rhs_parts.append(f"{sign} {term_str}")

    rhs = " ".join(rhs_parts)
    return f"{lhs} = {rhs}"


# Common queries

def get_formula_detail(conn, formula_id):
    return conn.execute(
        "SELECT *, json_extract(name, '$.en-us') AS name_en,"
        " science AS science_id,"
        " branch AS branch_id,"
        " subbranch AS subbranch_id,"
        " topic AS topic_id,"
        " json_extract(description, '$.en-us') AS description_en"
        " FROM formula WHERE id = ?",
        (formula_id,),
    ).fetchone()


def get_formula_items(conn, formula_id):
    return conn.execute("""
        SELECT fi.*, q.symbol AS quantity_symbol
        FROM formula_item fi
        LEFT JOIN quantity q ON q.id = fi.quantity_id
        WHERE fi.formula_id = ?
        ORDER BY fi.term, fi.is_primary DESC, fi.sort_order
    """, (formula_id,)).fetchall()


def get_formula_conditions(conn, formula_id):
    return conn.execute("""
        SELECT c.default_on, json_extract(c.name, '$.en-us') AS name_en,
               c.replacement_formula_id,
               json_extract(f2.name, '$.en-us') AS replacement_name
        FROM condition c
        JOIN formula f2 ON f2.id = c.replacement_formula_id
        WHERE c.formula_id = ?
        ORDER BY c.sort_order
    """, (formula_id,)).fetchall()


def get_formula_relations(conn, formula_id):
    return conn.execute("""
        SELECT fr.relation_type, fr.related_id,
               json_extract(f2.name, '$.en-us') AS related_name
        FROM formula_relation fr
        JOIN formula f2 ON f2.id = fr.related_id
        WHERE fr.formula_id = ?
        ORDER BY fr.relation_type
    """, (formula_id,)).fetchall()


def get_formula_quantities(conn, formula_id):
    rows = conn.execute("""
        SELECT DISTINCT q.id, q.name, q.symbol, json_extract(q.name, '$.en-us') AS name_en,
               COALESCE(json_extract(fi.quantity_name_overwrite, '$.en-us'), json_extract(q.name, '$.en-us')) AS display_name,
               q.default_unit, q.dim_M, q.dim_L, q.dim_T, q.dim_I, q.dim_\u0398, q.dim_N, q.dim_J
        FROM formula_item fi
        JOIN quantity q ON q.id = fi.quantity_id
        WHERE fi.formula_id = ?
    """, (formula_id,)).fetchall()
    return sort_quantities(rows)


def get_quantity_detail(conn, quantity_id):
    return conn.execute(
        "SELECT *, json_extract(name, '$.en-us') AS name_en,"
        " json_extract(description, '$.en-us') AS description_en"
        " FROM quantity WHERE id = ?",
        (quantity_id,),
    ).fetchone()


def get_quantity_units(conn, quantity_id):
    return conn.execute("""
        SELECT u.*, json_extract(u.name, '$.en-us') AS name_en
        FROM unit u WHERE u.quantity_id = ?
        ORDER BY u.default_unit DESC, u.unit_system
    """, (quantity_id,)).fetchall()


def get_quantity_formulas(conn, quantity_id):
    return conn.execute("""
        SELECT DISTINCT f.id, json_extract(f.name, '$.en-us') AS name_en,
               f.branch AS branch_id,
               f.subbranch AS subbranch_id,
               f.topic AS topic_id,
               f.difficulty
        FROM formula_item fi
        JOIN formula f ON f.id = fi.formula_id
        WHERE fi.quantity_id = ?
        ORDER BY f.branch, f.subbranch, f.topic, f.difficulty, f.id
    """, (quantity_id,)).fetchall()


def get_quantity_formulas_split(conn, quantity_id):
    """Return (primary_formulas, nonprimary_formulas) for a quantity."""
    rows = conn.execute("""
        SELECT DISTINCT f.id, json_extract(f.name, '$.en-us') AS name_en,
               f.branch AS branch_id,
               f.subbranch AS subbranch_id,
               f.topic AS topic_id,
               f.difficulty, fi.is_primary
        FROM formula_item fi
        JOIN formula f ON f.id = fi.formula_id
        WHERE fi.quantity_id = ?
        ORDER BY f.branch, f.subbranch, f.topic, f.difficulty, f.id
    """, (quantity_id,)).fetchall()
    primary = [r for r in rows if r["is_primary"]]
    nonprimary = [r for r in rows if not r["is_primary"]]
    return primary, nonprimary


def get_formula_primary_dimensions(conn, formula_id):
    """Compute dimensions from primary-term quantities."""
    items = conn.execute("""
        SELECT fi.var_exponent, q.dim_M, q.dim_L, q.dim_T,
               q.dim_I, q.dim_Θ, q.dim_N, q.dim_J
        FROM formula_item fi
        JOIN quantity q ON q.id = fi.quantity_id
        WHERE fi.formula_id = ? AND fi.is_primary = 1
    """, (formula_id,)).fetchall()
    dims = [0.0] * 7
    for r in items:
        exp = abs(r["var_exponent"] or 1)
        dims[0] += r["dim_M"] * exp
        dims[1] += r["dim_L"] * exp
        dims[2] += r["dim_T"] * exp
        dims[3] += r["dim_I"] * exp
        dims[4] += r["dim_Θ"] * exp
        dims[5] += r["dim_N"] * exp
        dims[6] += r["dim_J"] * exp
    return [int(d) for d in dims]


def get_si_unit_symbol(conn, quantity_id):
    """Get the SI base unit symbol for a quantity."""
    row = conn.execute("""
        SELECT u.symbol FROM unit u
        WHERE u.quantity_id = ? AND u.default_unit = 1
        LIMIT 1
    """, (quantity_id,)).fetchone()
    return row["symbol"] if row else ""


def get_quantity_related_formulas(conn, quantity_id):
    """Get formulas related via formula_relation that use this quantity."""
    return conn.execute("""
        SELECT DISTINCT f.id, json_extract(f.name, '$.en-us') AS name_en,
               fr.relation_type
        FROM formula_relation fr
        JOIN formula f ON f.id = fr.related_id
        JOIN formula_item fi ON fi.formula_id = f.id
        WHERE fi.quantity_id = ?
        ORDER BY fr.relation_type, f.id
    """, (quantity_id,)).fetchall()


def get_unit_detail(conn, unit_id):
    return conn.execute("""
        SELECT u.*, json_extract(q.name, '$.en-us') AS quantity_name,
               q.science AS science_id,
               q.branch AS branch_id,
               q.subbranch AS subbranch_id,
               q.topic AS topic_id,
               json_extract(u.name, '$.en-us') AS name_en
        FROM unit u JOIN quantity q ON q.id = u.quantity_id
        WHERE u.id = ?
    """, (unit_id,)).fetchone()


def search(conn, query, limit=30):
    """Full-text search across formulas, quantities, and units."""
    terms = " OR ".join(f'"{w}"*' for w in query.split())

    formulas = conn.execute("""
        SELECT 'formula' AS kind, fts.formula_id AS id,
               f.name AS name,
               json_extract(f.name, '$.en-us') AS name_en,
               f.branch AS branch_id,
               f.subbranch AS subbranch_id,
               f.topic AS topic_id,
               f.difficulty, NULL AS extra
        FROM formula_fts fts
        JOIN formula f ON f.id = fts.formula_id
        WHERE formula_fts MATCH ?
        ORDER BY rank
        LIMIT ?
    """, (terms, limit)).fetchall()

    quantities = conn.execute("""
        SELECT 'quantity' AS kind, fts.quantity_id AS id,
               q.name AS name,
               fts.name AS name_en, NULL, NULL, NULL,
               q.symbol AS extra
        FROM quantity_fts fts
        JOIN quantity q ON q.id = fts.quantity_id
        WHERE quantity_fts MATCH ?
        LIMIT ?
    """, (terms, limit)).fetchall()

    units = conn.execute("""
        SELECT 'unit' AS kind, fts.unit_id AS id,
               u.name AS name,
               fts.name AS name_en, NULL, NULL, NULL,
               fts.symbol AS extra
        FROM unit_fts fts
        JOIN unit u ON u.id = fts.unit_id
        WHERE unit_fts MATCH ?
        LIMIT ?
    """, (terms, limit)).fetchall()

    return list(formulas) + list(quantities) + list(units)


def get_dimension_symbol_maps(conn):
    """Return ({dim_sym: quantity_symbol}, {dim_sym: default_unit_symbol}, {dim_sym: dim_symbol}) for dimension display."""
    dim_var_ids = get_dim_var_ids(conn)
    DIM_SYMS = ["M", "L", "T", "I", "\u0398", "N", "J"]
    var_map = {}
    unit_map = {}
    dim_map = {}
    for dim_sym in DIM_SYMS:
        qid = dim_var_ids.get(dim_sym)
        if not qid:
            var_map[dim_sym] = dim_sym
            unit_map[dim_sym] = dim_sym
            dim_map[dim_sym] = dim_sym
            continue
        row = conn.execute("SELECT symbol, symbol_overwrite, default_unit FROM quantity WHERE id=?", (qid,)).fetchone()
        var_map[dim_sym] = row["symbol"] if row and row["symbol"] else dim_sym
        # unit mode: parse default_unit JSON to find the first unit's symbol
        unit_sym = None
        if row and row["default_unit"]:
            try:
                du = json.loads(row["default_unit"])
                if isinstance(du, list) and du:
                    uid = du[0].get("unit", "")
                    if uid:
                        urow = conn.execute("SELECT symbol FROM unit WHERE id=?", (uid,)).fetchone()
                        if urow:
                            unit_sym = urow["symbol"]
            except (ValueError, TypeError, IndexError):
                pass
        unit_map[dim_sym] = unit_sym if unit_sym else var_map[dim_sym]
        # dim mode: read symbol_overwrite["dim"] if available
        if qid:
            qrow = conn.execute("SELECT symbol_overwrite FROM quantity WHERE id=?", (qid,)).fetchone()
            if qrow and qrow["symbol_overwrite"]:
                try:
                    ow = json.loads(qrow["symbol_overwrite"])
                    dim_sym_val = ow.get("dim")
                    if dim_sym_val:
                        dim_map[dim_sym] = dim_sym_val
                        continue
                except (ValueError, TypeError):
                    pass
        dim_map[dim_sym] = dim_sym
    return var_map, unit_map, dim_map


def sort_quantities(quantities):
    """Sort quantities: 7 base dimensions in canonical order, rest alphabetically by id."""
    BASE_ORDER = {
        "mass": 1, "length": 2, "time": 3, "current": 4,
        "temperature": 5, "amount": 6, "luminous_intensity": 7,
    }
    def key(q):
        base = BASE_ORDER.get(q["id"], 99)
        return (0 if base < 99 else 1, base, q["id"])
    return sorted(quantities, key=key)


def get_all_quantities(conn):
    """Return all quantities with default_unit parsed and dimensions."""
    rows = conn.execute("""
        SELECT q.id, q.name, q.symbol,
               json_extract(q.name, '$.en-us') AS name_en,
               q.science AS science_id,
               q.branch AS branch_id,
               q.subbranch AS subbranch_id,
               q.topic AS topic_id,
               q.difficulty,
               q.is_dim,
               q.default_unit,
               q.dim_M, q.dim_L, q.dim_T, q.dim_I, q.dim_\u0398, q.dim_N, q.dim_J,
               (SELECT u.symbol FROM json_each(q.default_unit, '$') AS je
                  JOIN unit u ON u.id = je.value
                  WHERE u.default_unit = 1
                  ORDER BY je.rowid LIMIT 1) AS default_unit_symbol,
               (SELECT u.name FROM json_each(q.default_unit, '$') AS je
                  JOIN unit u ON u.id = je.value
                  WHERE u.default_unit = 1
                  ORDER BY je.rowid LIMIT 1) AS default_unit_name
        FROM quantity q
        ORDER BY q.id
    """).fetchall()
    return sort_quantities(rows)


def get_all_formulas(conn):
    """Return all formulas."""
    return conn.execute("""
        SELECT f.id, f.name, json_extract(f.name, '$.en-us') AS name_en,
               f.science AS science_id,
               f.branch, f.branch AS branch_id,
               f.subbranch, f.subbranch AS subbranch_id,
               f.topic, f.topic AS topic_id,
               f.difficulty
        FROM formula f
        ORDER BY f.science, f.branch, f.subbranch, f.topic, f.difficulty, f.id
    """).fetchall()


def get_formula_dimension_map(conn):
    """Return dict of {formula_id: {dim_M, dim_L, dim_T, dim_I, dim_Θ, dim_N, dim_J}}
    for ALL formulas, computed from primary-term quantities."""
    rows = conn.execute("""
        SELECT fi.formula_id,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_M) AS INTEGER) AS dim_M,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_L) AS INTEGER) AS dim_L,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_T) AS INTEGER) AS dim_T,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_I) AS INTEGER) AS dim_I,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_Θ) AS INTEGER) AS dim_Θ,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_N) AS INTEGER) AS dim_N,
               CAST(SUM(ABS(COALESCE(fi.var_exponent, 1)) * q.dim_J) AS INTEGER) AS dim_J
        FROM formula_item fi
        JOIN quantity q ON q.id = fi.quantity_id
        WHERE fi.is_primary = 1
        GROUP BY fi.formula_id
    """).fetchall()
    return {r["formula_id"]: {c: int(r[c] or 0) for c in DIM_COLS} for r in rows}


def get_formulas_containing_any_quantities(conn, qty_ids):
    """Return set of formula_ids that have formula_items for ANY of the given quantity IDs (OR)."""
    if not qty_ids:
        return None
    placeholders = ",".join("?" for _ in qty_ids)
    rows = conn.execute(f"""
        SELECT DISTINCT formula_id
        FROM formula_item
        WHERE quantity_id IN ({placeholders})
    """, qty_ids).fetchall()
    return {r["formula_id"] for r in rows}


def get_formulas_containing_all_quantities(conn, qty_ids):
    """Return set of formula_ids that have formula_items for ALL of the given quantity IDs (AND)."""
    if not qty_ids:
        return None
    n = len(qty_ids)
    placeholders = ",".join("?" for _ in qty_ids)
    rows = conn.execute(f"""
        SELECT formula_id, COUNT(DISTINCT quantity_id) as match_count
        FROM formula_item
        WHERE quantity_id IN ({placeholders})
        GROUP BY formula_id
        HAVING match_count = ?
    """, qty_ids + [n]).fetchall()
    return {r["formula_id"] for r in rows}


TABLE_ORDER = ["quantity", "unit", "formula", "formula_item", "condition", "formula_relation"]

TABLE_COLUMNS = {
    "formula": [
        "id", "name", "science", "branch", "subbranch", "topic", "difficulty",
        "description", "links",
    ],
    "formula_item": [
        "formula_id", "term", "is_primary", "sort_order",
        "coeff_value", "latex_coef", "coeff_exponent",
        "quantity_id", "var_exponent", "label",
        "symbol_overwrite", "quantity_name_overwrite",
        "latex_prefix", "latex_suffix",
    ],
    "condition": [
        "name", "formula_id", "replacement_formula_id", "default_on", "sort_order",
    ],
    "formula_relation": [
        "formula_id", "related_id", "relation_type",
    ],
    "quantity": [
        "id", "name", "symbol", "symbol_overwrite",
        "science", "branch", "subbranch", "topic",
        "difficulty", "description", "links", "is_dim", "default_unit",
        "dim_M", "dim_L", "dim_T", "dim_I", "dim_\u0398", "dim_N", "dim_J",
    ],
    "unit": [
        "id", "name", "symbol", "quantity_id", "default_unit", "unit_system",
        "factor", "latex_factor", "offset",
    ],
}


TABLE_INSERT = {
    "formula": """INSERT INTO formula
        (id, name, science, branch, subbranch, topic, difficulty, description, links)
        VALUES (?,?,?,?,?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name, science=excluded.science, branch=excluded.branch,
            subbranch=excluded.subbranch, topic=excluded.topic,
            difficulty=excluded.difficulty, description=excluded.description,
            links=excluded.links""",
    "formula_item": """INSERT INTO formula_item
        (formula_id, term, is_primary, sort_order, coeff_value, latex_coef,
         coeff_exponent, quantity_id, var_exponent, label,
         symbol_overwrite, quantity_name_overwrite,
         latex_prefix, latex_suffix)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(formula_id, term, is_primary, sort_order) DO UPDATE SET
            coeff_value=excluded.coeff_value, latex_coef=excluded.latex_coef,
            coeff_exponent=excluded.coeff_exponent, quantity_id=excluded.quantity_id,
            var_exponent=excluded.var_exponent, label=excluded.label,
            symbol_overwrite=excluded.symbol_overwrite,
            quantity_name_overwrite=excluded.quantity_name_overwrite,
            latex_prefix=excluded.latex_prefix, latex_suffix=excluded.latex_suffix""",
    "condition": """INSERT INTO condition
        (name, formula_id, replacement_formula_id, default_on, sort_order)
        VALUES (?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name, formula_id=excluded.formula_id,
            replacement_formula_id=excluded.replacement_formula_id,
            default_on=excluded.default_on, sort_order=excluded.sort_order""",
    "formula_relation": """INSERT INTO formula_relation
        (formula_id, related_id, relation_type)
        VALUES (?,?,?)
        ON CONFLICT(formula_id, related_id) DO UPDATE SET
            relation_type=excluded.relation_type""",
    "quantity": """INSERT INTO quantity
        (id, name, symbol, symbol_overwrite,
         science, branch, subbranch, topic,
         difficulty, description, links, is_dim, default_unit,
         dim_M, dim_L, dim_T, dim_I, dim_\u0398, dim_N, dim_J)
        VALUES (?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name, symbol=excluded.symbol,
            symbol_overwrite=excluded.symbol_overwrite,
            science=excluded.science,
            branch=excluded.branch, subbranch=excluded.subbranch,
            topic=excluded.topic, difficulty=excluded.difficulty,
            description=excluded.description, links=excluded.links,
            is_dim=excluded.is_dim, default_unit=excluded.default_unit,
            dim_M=excluded.dim_M, dim_L=excluded.dim_L, dim_T=excluded.dim_T,
            dim_I=excluded.dim_I, dim_\u0398=excluded.dim_\u0398,
            dim_N=excluded.dim_N, dim_J=excluded.dim_J""",
    "unit": """INSERT INTO unit
        (id, name, symbol, quantity_id, default_unit, unit_system,
         factor, latex_factor, offset)
        VALUES (?,?,?,?,?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name, symbol=excluded.symbol,
            quantity_id=excluded.quantity_id, default_unit=excluded.default_unit,
            unit_system=excluded.unit_system, factor=excluded.factor,
            latex_factor=excluded.latex_factor, offset=excluded.offset""",
}


def export_csv(conn):
    """Export all tables to a single CSV string with section headers."""
    buf = StringIO()
    for table in TABLE_ORDER:
        cols = TABLE_COLUMNS[table]
        buf.write(f"=== {table} ===\n")
        w = csv.writer(buf)
        w.writerow(cols)
        rows = conn.execute(f"SELECT {','.join(cols)} FROM {table} ORDER BY rowid").fetchall()
        for r in rows:
            w.writerow([r[c] for c in cols])
        buf.write("\n")
    return buf.getvalue()


def _table_rows(conn):
    """Yield (table_name, headers, rows) for all tables."""
    for table in TABLE_ORDER:
        cols = TABLE_COLUMNS[table]
        rows = conn.execute(f"SELECT {','.join(cols)} FROM {table} ORDER BY rowid").fetchall()
        yield table, cols, [[r[c] for c in cols] for r in rows]


def export_csv_dir(conn, directory):
    """Export each table to a separate CSV file in directory."""
    p = Path(directory)
    p.mkdir(parents=True, exist_ok=True)
    for table, cols, rows in _table_rows(conn):
        path = p / f"{table}.csv"
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(cols)
            w.writerows(rows)


def export_xlsx(conn, path):
    """Export all tables as sheets in an XLSX workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    first = True
    for table, cols, rows in _table_rows(conn):
        if first:
            ws = wb.active
            ws.title = table[:31]
            first = False
        else:
            ws = wb.create_sheet(title=table[:31])
        ws.append(cols)
        for row in rows:
            ws.append(row)
    wb.save(path)


def export_ods(conn, path):
    """Export all tables as sheets in an ODS spreadsheet."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    doc = OpenDocumentSpreadsheet()
    for table, cols, rows in _table_rows(conn):
        tbl = Table(name=table[:31])
        doc.spreadsheet.addElement(tbl)
        for row_data in [cols] + rows:
            tr = TableRow()
            for val in row_data:
                tc = TableCell()
                p = P(text=str(val) if val is not None else "")
                tc.addElement(p)
                tr.addElement(tc)
            tbl.addElement(tr)
    doc.save(path)


_NOT_NULL_TEXT = {
    "formula": {"id", "name"},
    "formula_item": {"formula_id"},
    "condition": {"formula_id", "replacement_formula_id"},
    "formula_relation": {"formula_id", "related_id", "relation_type"},
    "quantity": {"id", "name", "symbol"},
    "unit": {"id", "name", "symbol", "quantity_id"},
}


def _insert_rows(conn, table, headers, rows):
    """Insert rows from one table, updating counts dict."""
    expected = TABLE_COLUMNS[table]
    col_indices = [headers.index(col) for col in expected if col in headers]
    if not col_indices:
        return 0
    sql = TABLE_INSERT[table]
    nn_cols = _NOT_NULL_TEXT.get(table, set())
    nn_indices = {expected.index(c) for c in nn_cols if c in expected}
    cleaned = []
    for row in rows:
        if not row:
            continue
        vals = [row[i] if i < len(row) else "" for i in col_indices]
        cleaned.append(tuple(
            v if (v != "" or i in nn_indices) else None
            for i, v in enumerate(vals)
        ))
    if cleaned:
        conn.executemany(sql, cleaned)
    return len(cleaned)



def import_csv_dir(conn, directory):
    """Import tables from per-table CSV files in a directory."""
    p = Path(directory)
    return _import_worksheets(conn, {
        table: _read_csv(p / f"{table}.csv")
        for table in TABLE_ORDER
        if (p / f"{table}.csv").exists()
    })


def _read_csv(path):
    """Read a CSV file, returning (headers, rows)."""
    with open(path, newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        rows = list(reader)
    return headers, rows


def import_xlsx(conn, path):
    """Import tables from XLSX workbook sheets."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            sheets[name] = (list(rows[0]), [list(r) for r in rows[1:]])
    wb.close()
    return _import_worksheets(conn, sheets)


def import_ods(conn, path):
    """Import tables from ODS spreadsheet sheets."""
    from odf.opendocument import load
    from odf.table import Table, TableRow
    from odf.text import P
    doc = load(path)
    sheets = {}
    for table_elem in doc.getElementsByType(Table):
        name = table_elem.getAttribute("name")
        rows = []
        for row_elem in table_elem.getElementsByType(TableRow):
            cells = []
            for cell in row_elem.childNodes:
                cell_text = None
                for p in cell.getElementsByType(P):
                    texts = []
                    for n in p.childNodes:
                        if hasattr(n, "data"):
                            texts.append(n.data)
                    cell_text = "".join(texts)
                    break
                cells.append(cell_text)
            if any(c is not None and c.strip() for c in cells):
                rows.append(cells)
        if rows:
            sheets[name] = (list(rows[0]), rows[1:])
    return _import_worksheets(conn, sheets)


def _import_worksheets(conn, sheets):
    """Import from a dict of {table_name: (headers, rows)}. Returns counts."""
    counts = {}
    table_map = {}
    for name in sheets:
        for t in TABLE_ORDER:
            if t.startswith(name) or name.startswith(t):
                table_map[t] = sheets[name]
                break
    try:
        conn.execute("BEGIN")
        for table in TABLE_ORDER:
            if table in table_map:
                headers, rows = table_map[table]
                n = _insert_rows(conn, table, headers, rows)
                counts[table] = n
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return counts


def import_csv(conn, csv_str):
    """Import tables from a CSV string with section headers. Returns counts."""
    counts = {}
    reader = csv.reader(StringIO(csv_str))
    current_table = None
    header_row = None
    col_indices = []
    nn_indices = set()
    rows_buffer = []

    def flush():
        if not current_table or not rows_buffer:
            return
        sql = TABLE_INSERT[current_table]
        cleaned = []
        for row in rows_buffer:
            vals = [row[i] if i < len(row) else "" for i in col_indices]
            cleaned.append(tuple(
                v if (v != "" or i in nn_indices) else None
                for i, v in enumerate(vals)
            ))
        conn.executemany(sql, cleaned)
        counts[current_table] = len(rows_buffer)
        rows_buffer.clear()

    try:
        conn.execute("BEGIN")
        for row in reader:
            if not row or all(c.strip() == "" for c in row):
                continue
            if row[0].startswith("=== ") and row[0].endswith(" ==="):
                flush()
                current_table = row[0][4:-4].strip()
                header_row = None
                col_indices = []
                nn_indices = set()
                continue
            if current_table and header_row is None:
                header_row = row
                expected = TABLE_COLUMNS.get(current_table, [])
                col_indices = [header_row.index(col) for col in expected if col in header_row]
                nn_cols = _NOT_NULL_TEXT.get(current_table, set())
                nn_indices = {
                    idx for idx, col in enumerate(expected)
                    if col in nn_cols and col in header_row
                }
                continue
            if current_table and col_indices:
                rows_buffer.append(tuple(row))

        flush()
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return counts
